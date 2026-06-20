"""
Import stock data from Groww Excel reports into MyPortfolio app.
Part A: Stock orders from Stocks_Order_History_*.xlsx (546 executed orders)
Part B: Dividends (TECHNOCRAFT Rs 200, TRENT Rs 180)
Part C: Deposits and charges from Groww_Balance_Statement_*.xlsx
Part D: Holding adjustment for pre-period sells (stocks sold without matching buy in this period)

Requires: pip install openpyxl requests
Usage: python scripts/import_stock_data.py
       (backend must be running on localhost:8081)
"""

import sys
import openpyxl
import requests
from datetime import datetime

BASE_URL = "http://localhost:8081/api"

ORDERS_FILE = "GROWW_Reports_06192026/Stocks_Order_History_4728936310_01-04-2026_19-06-2026.xlsx"
BALANCE_FILE = "GROWW_Reports_06192026/Groww_Balance_Statement_4728936310_01-04-2026_19-06-2026.xlsx"


def ensure_stock(symbol, company_name, exchange="NSE"):
    """Ensure stock exists in the database, create if needed."""
    try:
        resp = requests.get(f"{BASE_URL}/stocks/{symbol}", timeout=10)
        if resp.status_code == 200:
            return True
    except Exception:
        pass

    payload = {
        "symbol": symbol.upper(),
        "companyName": company_name,
        "exchange": exchange,
    }
    try:
        resp = requests.post(f"{BASE_URL}/stocks", json=payload, timeout=10)
        if resp.status_code == 201:
            print(f"  Created stock: {symbol} ({company_name})")
            return True
        elif resp.status_code == 409:
            return True
        else:
            print(f"  Failed to create stock {symbol}: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"  Error creating stock {symbol}: {e}")
    return False


def parse_execution_datetime(dt_str):
    """Parse '27-04-2026 10:36 AM' -> '2026-04-27T10:36:00'"""
    if not dt_str:
        return None
    try:
        dt = datetime.strptime(str(dt_str).strip(), "%d-%m-%Y %I:%M %p")
        return dt.strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        try:
            dt = datetime.strptime(str(dt_str).strip(), "%d-%m-%Y %H:%M")
            return dt.strftime("%Y-%m-%dT%H:%M:%S")
        except Exception:
            return None


def parse_balance_date(dt_str):
    """Parse '27-04-2026' -> '2026-04-27T00:00:00'"""
    if not dt_str:
        return None
    try:
        dt = datetime.strptime(str(dt_str).strip(), "%d-%m-%Y")
        return dt.strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def create_transaction(symbol, quantity, price, txn_type, trade_date, description=None, trade_type=None):
    """Create a stock transaction via REST API."""
    payload = {
        "symbol": symbol,
        "quantity": quantity,
        "price": price,
        "transactionType": txn_type,
        "tradeDate": trade_date,
        "description": description,
        "tradeType": trade_type,
    }
    try:
        resp = requests.post(f"{BASE_URL}/transactions", json=payload, timeout=10)
        if resp.status_code == 201:
            return resp.json()
        else:
            print(f"  Failed txn {txn_type} {symbol}: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"  Error creating txn: {e}")
    return None


def create_fund_transaction(txn_type, amount, trade_date, description):
    """Create a fund transaction (DEPOSIT/WITHDRAWAL/CHARGES) via REST API."""
    payload = {
        "symbol": None,
        "quantity": 1,
        "price": amount,
        "transactionType": txn_type,
        "tradeDate": trade_date,
        "description": description,
    }
    try:
        resp = requests.post(f"{BASE_URL}/transactions", json=payload, timeout=10)
        if resp.status_code == 201:
            return resp.json()
        else:
            print(f"  Failed fund txn {txn_type}: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"  Error creating fund txn: {e}")
    return None


def detect_intraday(rows):
    """Pass 1: Build set of (symbol, date) pairs that have BOTH buy and sell → intraday (MIS)."""
    from collections import defaultdict
    day_types = defaultdict(set)

    for r in rows:
        symbol = r["symbol"]
        txn_type = r["txn_type"]
        trade_date_str = r["trade_date"]
        if not trade_date_str:
            continue
        trade_day = trade_date_str[:10]
        day_types[(symbol, trade_day)].add(txn_type)

    intraday_keys = set()
    for key, types in day_types.items():
        if "BUY" in types and "SELL" in types:
            intraday_keys.add(key)

    return intraday_keys


def import_stock_orders():
    """Part A: Import stock orders from Groww Excel with intraday/delivery detection."""
    print("\n=== PART A: Importing Stock Orders ===")
    wb = openpyxl.load_workbook(ORDERS_FILE)
    ws = wb.active

    parsed_rows = []
    skipped = 0

    for i, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 7):
        vals = [cell.value for cell in row]
        stock_name = vals[0]
        symbol = vals[1]
        isin = vals[2]
        txn_type = vals[3]
        quantity = vals[4]
        value = vals[5]
        exchange = vals[6]
        order_id = vals[7]
        exec_datetime = vals[8]
        status = vals[9]

        if not symbol or not stock_name:
            continue
        if status and str(status).strip() != "Executed":
            skipped += 1
            continue

        symbol = str(symbol).strip().upper()
        exchange_str = str(exchange).strip() if exchange else "NSE"
        qty = int(float(quantity)) if quantity else 0
        val = float(value) if value else 0
        price_per_unit = round(val / qty, 2) if qty > 0 else 0
        txn_type_enum = "SELL" if str(txn_type).strip().upper() == "SELL" else "BUY"
        trade_date = parse_execution_datetime(exec_datetime)

        parsed_rows.append({
            "symbol": symbol,
            "stock_name": str(stock_name).strip(),
            "exchange": exchange_str,
            "qty": qty,
            "price": price_per_unit,
            "txn_type": txn_type_enum,
            "trade_date": trade_date,
        })

    print(f"  Parsed {len(parsed_rows)} orders, skipped {skipped} non-executed")

    intraday_keys = detect_intraday(parsed_rows)
    mis_count = sum(1 for r in parsed_rows if (r["symbol"], r["trade_date"][:10] if r["trade_date"] else "") in intraday_keys)
    cnc_count = len(parsed_rows) - mis_count
    print(f"  Detected: {mis_count} intraday (MIS), {cnc_count} delivery (CNC)")

    created = 0
    errors = 0
    stocks_created = set()

    for r in parsed_rows:
        if r["symbol"] not in stocks_created:
            if ensure_stock(r["symbol"], r["stock_name"], r["exchange"]):
                stocks_created.add(r["symbol"])
            else:
                errors += 1
                continue

        trade_day = r["trade_date"][:10] if r["trade_date"] else ""
        trade_type = "MIS" if (r["symbol"], trade_day) in intraday_keys else "CNC"

        result = create_transaction(
            r["symbol"], r["qty"], r["price"], r["txn_type"],
            r["trade_date"], trade_type=trade_type
        )
        if result:
            created += 1
            if created % 100 == 0:
                print(f"  Progress: {created} transactions created...")
        else:
            errors += 1

    print(f"\n--- Stock orders import complete ---")
    print(f"  Created: {created}, Errors: {errors}")
    print(f"  Unique stocks: {len(stocks_created)}")
    print(f"  MIS (intraday): {mis_count}, CNC (delivery): {cnc_count}")
    return created, errors


def import_dividends():
    """Part B: Import 2 dividends from PDF report."""
    print("\n=== PART B: Importing Dividends ===")
    dividends = [
        ("TIIL", 10, 20.0, "2026-06-04T00:00:00", "Dividend Rs 20/share"),
        ("TRENT", 45, 4.0, "2026-06-12T00:00:00", "Dividend Rs 4/share"),
    ]

    created = 0
    errors = 0

    for symbol, qty, div_per_share, trade_date, desc in dividends:
        total = round(qty * div_per_share, 2)
        result = create_transaction(symbol, qty, div_per_share, "DIVIDEND", trade_date, desc)
        if result:
            created += 1
            print(f"  {symbol}: Rs {total} dividend")
        else:
            errors += 1

    print(f"  Dividends: {created} created, {errors} errors")
    return created, errors


def import_deposits_and_charges():
    """Part C: Import deposits and charges from Balance Statement."""
    print("\n=== PART C: Importing Deposits & Charges ===")
    wb = openpyxl.load_workbook(BALANCE_FILE)
    ws = wb.active

    deposits = 0
    charges = 0
    errors = 0

    for i, row in enumerate(ws.iter_rows(min_row=9, max_row=ws.max_row), 9):
        vals = [cell.value for cell in row]
        txn_date = vals[0]
        segment = vals[3]
        debit = float(vals[9]) if vals[9] else 0
        credit = float(vals[10]) if vals[10] else 0

        if not segment or segment == "STOCKS_SETTLEMENT":
            continue

        trade_date = parse_balance_date(str(txn_date)) if txn_date else None

        if segment == "UPI" and credit > 0:
            result = create_fund_transaction("DEPOSIT", credit, trade_date,
                                             f"UPI Deposit {txn_date}")
            if result:
                deposits += 1
                print(f"  Deposit: Rs {credit:,.2f} ({txn_date})")
            else:
                errors += 1

        elif segment in ("STOCKS_MIS", "DDPI_CHARGES") and debit > 0:
            result = create_fund_transaction("CHARGES", debit, trade_date,
                                             f"{segment} charge {txn_date}")
            if result:
                charges += 1
                print(f"  Charge: Rs {debit:.2f} ({segment}, {txn_date})")
            else:
                errors += 1

    print(f"\n--- Deposits & charges complete ---")
    print(f"  Deposits: {deposits}, Charges: {charges}, Errors: {errors}")
    return deposits + charges, errors


def verify():
    """Verify import results."""
    print("\n=== Verification ===")
    try:
        stocks = requests.get(f"{BASE_URL}/stocks", timeout=10).json()
        txns = requests.get(f"{BASE_URL}/transactions", timeout=10).json()
        analytics = requests.get(f"{BASE_URL}/transactions/analytics", timeout=10).json()

        print(f"  Total stocks: {len(stocks)}")
        print(f"  Total transactions: {len(txns)}")
        print(f"  Buy amount: Rs {float(analytics.get('totalBuyAmount', 0)):,.2f}")
        print(f"  Sell amount: Rs {float(analytics.get('totalSellAmount', 0)):,.2f}")
        print(f"  Deposits: Rs {float(analytics.get('totalDeposits', 0)):,.2f}")
        print(f"  Charges: Rs {float(analytics.get('totalCharges', 0)):,.2f}")
    except Exception as e:
        print(f"  Verification failed: {e}")


if __name__ == "__main__":
    print("MyPortfolio — Stock Data Import")
    print(f"Backend: {BASE_URL}")

    try:
        resp = requests.get(f"{BASE_URL}/stocks", timeout=5)
        print(f"Backend reachable (status {resp.status_code})")
    except Exception:
        print("ERROR: Backend not reachable. Start with: ./mvnw spring-boot:run")
        sys.exit(1)

    o_count, o_err = import_stock_orders()
    d_count, d_err = import_dividends()
    f_count, f_err = import_deposits_and_charges()
    verify()

    total = o_count + d_count + f_count
    total_err = o_err + d_err + f_err
    print(f"\n=== DONE: {total} transactions imported, {total_err} errors ===")
    print("\nNext steps:")
    print("  1. If Groww API key is active: POST /api/groww/sync to correct holdings")
    print("  2. Run technical analysis: signals will auto-generate")

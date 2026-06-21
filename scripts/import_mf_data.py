"""
Import Mutual Fund data from Groww Excel reports into MyPortfolio app.
Part A: 12 MF holdings from Mutual_Funds_*.xlsx
Part B: 142 MF capital gains trades from MF_Capital_Gains_2025-2026.xlsx

Requires: pip install openpyxl requests
Usage: python scripts/import_mf_data.py
       (backend must be running on localhost:8081 with V11-V13 applied)
"""

import sys
import time
import openpyxl
import requests
from datetime import datetime

BASE_URL = "http://localhost:8081/api"
AUTH_HEADERS = {}

def login(email="sampath12082@gmail.com", password="Admin@123"):
    global AUTH_HEADERS
    resp = requests.post(f"{BASE_URL}/auth/login", json={"email": email, "password": password}, timeout=10)
    if resp.status_code == 200:
        AUTH_HEADERS = {"Authorization": f"Bearer {resp.json()['accessToken']}"}
        print(f"Logged in as {email}")
    else:
        print(f"Login failed: {resp.status_code}")
        sys.exit(1)

HOLDINGS_FILE = "GROWW_Reports_06192026/Mutual_Funds_4728936310_20-06-2026_20-06-2026.xlsx"
CAPITAL_GAINS_FILE = "GROWW_Reports_06192026/Mutual_Funds_Capital_Gains_Report_01-04-2025_31-03-2026.xlsx"

# Scheme codes from capital gains report (scheme_name -> scheme_code)
KNOWN_SCHEME_CODES = {
    "Baroda BNP Paribas Energy Opportunities Fund Growth": "153226",
    "Franklin India Flexi Cap Fund Growth": "100520",
    "HDFC Balanced Advantage Fund Direct Growth": "118968",
    "HDFC Flexi Cap Fund Growth": "101762",
    "HDFC Focused Fund Growth": "102760",
    "HDFC Mid Cap Fund Growth": "105758",
    "ICICI Prudential Value Fund Growth": "102594",
    "Nippon India Small Cap Fund Growth": "113177",
    "PGIM India Midcap Fund Growth": "125305",
    "Parag Parikh Flexi Cap Fund Growth": "122640",
    "Quant Small Cap Fund Growth": "100177",
    "SBI Contra Fund Growth": "102414",
    "HDFC Gold ETF Fund of Fund Direct Plan Growth": "119132",
    "HDFC Gold ETF Fund of Fund Growth": "115934",
    "HDFC ELSS Tax Saver Fund Growth": "101979",
}


def search_amfi(query):
    """Search AMFI via the app's API to find scheme code."""
    try:
        resp = requests.get(f"{BASE_URL}/mf/funds/search", params={"query": query}, headers=AUTH_HEADERS, timeout=30)
        if resp.status_code == 200:
            results = resp.json()
            if results:
                return results[0]
    except Exception as e:
        print(f"  AMFI search failed for '{query}': {e}")
    return None


def find_scheme_code(scheme_name):
    """Find scheme code from known mapping or AMFI search."""
    if scheme_name in KNOWN_SCHEME_CODES:
        return KNOWN_SCHEME_CODES[scheme_name]

    result = search_amfi(scheme_name)
    if result:
        return str(result.get("schemeCode", ""))
    return None


def create_fund(scheme_code, scheme_name, fund_house, category, fund_type=None):
    """Create a mutual fund via REST API. Skips if already exists."""
    payload = {
        "schemeCode": str(scheme_code),
        "schemeName": scheme_name,
        "fundHouse": fund_house,
        "category": category,
        "fundType": fund_type or "Open Ended",
    }
    try:
        resp = requests.post(f"{BASE_URL}/mf/funds", json=payload, headers=AUTH_HEADERS, timeout=10)
        if resp.status_code == 201:
            print(f"  Created fund: {scheme_name}")
            return resp.json()
        elif resp.status_code == 409:
            print(f"  Fund exists: {scheme_name}")
            return {"schemeCode": str(scheme_code)}
        else:
            print(f"  Failed to create fund {scheme_name}: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"  Error creating fund: {e}")
    return None


def create_holding(scheme_code, units, avg_nav):
    """Create/update a holding via REST API."""
    payload = {
        "schemeCode": str(scheme_code),
        "units": units,
        "averageNav": avg_nav,
    }
    try:
        resp = requests.post(f"{BASE_URL}/mf/holdings", json=payload, headers=AUTH_HEADERS, timeout=10)
        if resp.status_code == 201:
            print(f"  Created holding: {scheme_code} ({units} units @ {avg_nav})")
            return resp.json()
        else:
            print(f"  Failed to create holding {scheme_code}: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"  Error creating holding: {e}")
    return None


def create_mf_transaction(scheme_code, units, nav, amount, txn_type, trade_date, folio, description=None):
    """Create a MF transaction via REST API."""
    payload = {
        "schemeCode": str(scheme_code),
        "units": units,
        "nav": nav,
        "amount": amount,
        "transactionType": txn_type,
        "tradeDate": trade_date,
        "folioNumber": str(folio) if folio else None,
        "description": description,
    }
    try:
        resp = requests.post(f"{BASE_URL}/mf/transactions", json=payload, headers=AUTH_HEADERS, timeout=10)
        if resp.status_code == 201:
            return resp.json()
        else:
            print(f"  Failed to create txn {scheme_code} {txn_type}: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"  Error creating txn: {e}")
    return None


def import_holdings():
    """Part A: Import 12 MF holdings from Groww Excel."""
    print("\n=== PART A: Importing MF Holdings ===")
    wb = openpyxl.load_workbook(HOLDINGS_FILE)
    ws = wb.active

    created = 0
    errors = 0

    for i, row in enumerate(ws.iter_rows(min_row=23, max_row=34), 23):
        vals = [cell.value for cell in row]
        scheme_name = vals[0]
        if not scheme_name:
            continue

        amc = vals[1]
        category = vals[2]
        sub_category = vals[3]
        folio = vals[4]
        units = float(vals[6])
        invested_str = vals[7]
        invested = float(invested_str) if invested_str else 0.0

        full_category = f"{category}/{sub_category}" if sub_category else category

        print(f"\nRow {i}: {scheme_name}")
        print(f"  AMC={amc}, Category={full_category}, Units={units}, Invested={invested}")

        scheme_code = find_scheme_code(scheme_name)
        if not scheme_code:
            print(f"  ERROR: Could not find scheme code for {scheme_name}")
            errors += 1
            continue

        avg_nav = round(invested / units, 4) if units > 0 else 0
        print(f"  Scheme code={scheme_code}, Avg NAV={avg_nav}")

        fund = create_fund(scheme_code, scheme_name, amc, full_category)
        if not fund:
            errors += 1
            continue

        holding = create_holding(scheme_code, round(units, 4), round(avg_nav, 4))
        if holding:
            created += 1
        else:
            errors += 1

    print(f"\n--- Holdings import complete: {created} created, {errors} errors ---")
    return created, errors


def import_capital_gains():
    """Part B: Import 142 MF capital gains trades from Groww Excel."""
    print("\n=== PART B: Importing MF Capital Gains Trades ===")
    wb = openpyxl.load_workbook(CAPITAL_GAINS_FILE)
    ws = wb.active

    purchases = 0
    redemptions = 0
    funds_created = 0
    errors = 0
    seen_purchase_ids = set()
    seen_redeem_ids = set()

    existing_funds = set()
    try:
        resp = requests.get(f"{BASE_URL}/mf/funds", headers=AUTH_HEADERS, timeout=10)
        if resp.status_code == 200:
            for f in resp.json():
                existing_funds.add(str(f.get("schemeCode", "")))
    except Exception:
        pass

    for i, row in enumerate(ws.iter_rows(min_row=22, max_row=162), 22):
        vals = [cell.value for cell in row]
        scheme_name = vals[0]
        scheme_code = vals[1]

        if not scheme_name or not scheme_code:
            continue

        scheme_code = str(scheme_code)
        if scheme_code == "Scheme Code":
            continue

        category = vals[2]
        folio = vals[3]
        purchase_txn_id = vals[4]
        purchase_date = vals[5]
        matched_qty = float(vals[6]) if vals[6] else 0
        purchase_price = float(vals[7]) if vals[7] else 0
        redeem_txn_id = vals[8]
        redeem_date = vals[9]
        redeem_price = float(vals[11]) if vals[11] else 0

        if scheme_code not in existing_funds:
            fund = create_fund(scheme_code, scheme_name, None, category)
            if fund:
                existing_funds.add(scheme_code)
                funds_created += 1
            else:
                errors += 1
                continue

        purchase_amount = round(matched_qty * purchase_price, 2)
        redeem_amount = round(matched_qty * redeem_price, 2)

        trade_date_buy = f"{purchase_date}T00:00:00" if purchase_date else None
        trade_date_sell = f"{redeem_date}T00:00:00" if redeem_date else None

        dedup_buy_key = f"{purchase_txn_id}_{matched_qty}_{purchase_price}"
        if dedup_buy_key not in seen_purchase_ids:
            seen_purchase_ids.add(dedup_buy_key)
            result = create_mf_transaction(
                scheme_code, round(matched_qty, 4), round(purchase_price, 4),
                purchase_amount, "PURCHASE", trade_date_buy, folio,
                f"CG Import: {purchase_txn_id}"
            )
            if result:
                purchases += 1
            else:
                errors += 1

        dedup_sell_key = f"{redeem_txn_id}_{matched_qty}_{redeem_price}"
        if dedup_sell_key not in seen_redeem_ids:
            seen_redeem_ids.add(dedup_sell_key)
            result = create_mf_transaction(
                scheme_code, round(matched_qty, 4), round(redeem_price, 4),
                redeem_amount, "REDEMPTION", trade_date_sell, folio,
                f"CG Import: {redeem_txn_id}"
            )
            if result:
                redemptions += 1
            else:
                errors += 1

    print(f"\n--- Capital gains import complete ---")
    print(f"  Funds created: {funds_created}")
    print(f"  Purchases: {purchases}")
    print(f"  Redemptions: {redemptions}")
    print(f"  Errors: {errors}")
    return purchases + redemptions, errors


def refresh_navs():
    """Refresh NAVs from AMFI feed."""
    print("\n=== Refreshing NAVs from AMFI ===")
    try:
        resp = requests.post(f"{BASE_URL}/mf/funds/refresh-nav", headers=AUTH_HEADERS, timeout=60)
        if resp.status_code == 200:
            print("  NAVs refreshed successfully")
        else:
            print(f"  NAV refresh returned: {resp.status_code}")
    except Exception as e:
        print(f"  NAV refresh failed: {e}")


def verify():
    """Verify import results."""
    print("\n=== Verification ===")
    try:
        funds = requests.get(f"{BASE_URL}/mf/funds", headers=AUTH_HEADERS, timeout=10).json()
        holdings = requests.get(f"{BASE_URL}/mf/holdings", headers=AUTH_HEADERS, timeout=10).json()
        txns = requests.get(f"{BASE_URL}/mf/transactions", headers=AUTH_HEADERS, timeout=10).json()
        print(f"  Funds: {len(funds)}")
        print(f"  Holdings (active): {len(holdings)}")
        print(f"  Transactions: {len(txns)}")

        if holdings:
            total_invested = sum(float(h.get("investedAmount", 0) or 0) for h in holdings)
            total_current = sum(float(h.get("currentValue", 0) or 0) for h in holdings)
            print(f"  Total invested: Rs {total_invested:,.2f}")
            print(f"  Total current: Rs {total_current:,.2f}")
    except Exception as e:
        print(f"  Verification failed: {e}")


if __name__ == "__main__":
    print("MyPortfolio — MF Data Import")
    print(f"Backend: {BASE_URL}")

    login()

    try:
        resp = requests.get(f"{BASE_URL}/mf/funds", headers=AUTH_HEADERS, timeout=5)
        print(f"Backend reachable (status {resp.status_code})")
    except Exception:
        print("ERROR: Backend not reachable. Start with: ./mvnw spring-boot:run")
        sys.exit(1)

    h_count, h_err = import_holdings()
    t_count, t_err = import_capital_gains()
    refresh_navs()
    verify()

    print(f"\n=== DONE: {h_count} holdings, {t_count} transactions, {h_err + t_err} errors ===")
